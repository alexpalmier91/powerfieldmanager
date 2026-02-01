# app/routers/superuser_agent_orders_import.py
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import load_workbook

from app.core.security import get_current_user
from app.db.session import get_async_session
from app.db.models import (
    Agent,
    Client,
    Labo,
    LaboClient,
    Order,
    OrderItem,
    Product,          # 👈 ajouté
    UserRole,
    labo_agent,
)

router = APIRouter(
    prefix="/api-zenhub/superuser",
    tags=["superuser-agent-orders-import"],
)


# ============================================================
#   Schemas
# ============================================================

class ImportOrderSummary(BaseModel):
    order_number: str
    client_id: Optional[int] = None
    agent_id: Optional[int] = None
    total_ht: float


class ImportResult(BaseModel):
    labo_id: int
    total_rows: int
    created_orders: int
    updated_orders: int
    skipped_rows: int
    errors: List[str]
    orders: List[ImportOrderSummary]


# ============================================================
#   Helpers parsing
# ============================================================

def parse_date_safe(value: Any) -> Optional[date]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def parse_decimal_safe(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).replace(",", ".").strip()
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def normalize_agent_key(raw: Any) -> Optional[str]:
    if not raw:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    return s


def normalize_code_client(v: Any) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().lower()
    return s if s else None


def normalize_sku(v: Any) -> Optional[str]:
    """Normalisation SKU pour le mapping produit."""
    if not v:
        return None
    s = str(v).strip()
    return s.upper() if s else None


# ============================================================
#   DB helpers
# ============================================================

async def load_labo(labo_id: int, session: AsyncSession) -> Labo:
    res = await session.execute(select(Labo).where(Labo.id == labo_id))
    labo = res.scalar_one_or_none()
    if not labo:
        raise HTTPException(status_code=404, detail=f"Labo {labo_id} not found")
    return labo


async def build_agent_index_for_labo(labo_id: int, session: AsyncSession) -> Dict[str, Agent]:
    res = await session.execute(
        select(Agent)
        .join(labo_agent, labo_agent.c.agent_id == Agent.id)
        .where(labo_agent.c.labo_id == labo_id)
    )
    agents = res.scalars().all()

    index: Dict[str, Agent] = {}
    for ag in agents:
        ln = (ag.lastname or "").strip().lower()
        fn = (ag.firstname or "").strip().lower()

        if ln:
            index[ln] = ag
        if fn and ln:
            index[f"{fn} {ln}"] = ag

    return index


async def build_client_index_for_labo(labo_id: int, session: AsyncSession) -> Dict[str, Client]:
    res = await session.execute(
        select(LaboClient, Client)
        .join(Client, Client.id == LaboClient.client_id)
        .where(LaboClient.labo_id == labo_id)
    )
    rows = res.all()

    index: Dict[str, Client] = {}
    for lc, c in rows:
        if lc.code_client:
            key = lc.code_client.strip().lower()
            if key:
                index[key] = c
    return index




async def build_product_index_for_labo(labo_id: int, session: AsyncSession) -> Dict[str, tuple[int, Decimal]]:
    """
    Mapping SKU normalisé -> (product_id, vat_rate) pour le labo.
    Permet de remplir OrderItem.product_id et de calculer la TVA à l'import.
    """
    res = await session.execute(
        select(Product.id, Product.sku, Product.vat_rate).where(Product.labo_id == labo_id)
    )
    rows = res.all()

    index: Dict[str, tuple[int, Decimal]] = {}
    for pid, sku, vat_rate in rows:
        key = normalize_sku(sku)
        if key:
            index[key] = (pid, Decimal(vat_rate or 0))
    return index



# ============================================================
#   Superuser Auth
# ============================================================

def get_current_superuser(current_user: Any = Depends(get_current_user)) -> Any:
    role = getattr(current_user, "role", None)

    if role is None and isinstance(current_user, dict):
        role = current_user.get("role")

    if isinstance(role, str):
        try:
            role_enum = UserRole(role)
        except Exception:
            role_enum = None
    else:
        role_enum = role

    if role_enum not in (UserRole.SUPERUSER, UserRole.SUPERADMIN):
        raise HTTPException(status_code=403, detail="Forbidden")

    return current_user


# ============================================================
#   ENDPOINT IMPORT
# ============================================================

@router.post(
    "/agent-orders/import",
    response_model=ImportResult,
)
async def import_agent_orders_from_excel(
    labo_id: int = Query(...),
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    current_user: Any = Depends(get_current_superuser),
):
    labo = await load_labo(labo_id, session)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel: {e}")

    ws = wb["merged"] if "merged" in wb.sheetnames else wb.worksheets[0]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for i, name in enumerate(header):
        if name:
            col_index[str(name).strip().lower()] = i

    required = [
        "order_number", "company_name", "order_date", "delivery_date",
        "agent", "sku", "qty", "unit_ht", "code_client"
    ]
    miss = [c for c in required if c not in col_index]
    if miss:
        raise HTTPException(status_code=400, detail=f"Missing columns: {miss}")

    has_comment = "comment" in col_index

    agent_index = await build_agent_index_for_labo(labo_id, session)
    client_index = await build_client_index_for_labo(labo_id, session)
    product_index = await build_product_index_for_labo(labo_id, session)  # 👈 nouveau

    total_rows = 0
    skipped_rows = 0
    errors: List[str] = []
    grouped: dict[str, list[dict]] = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1

        def get(col: str):
            i = col_index.get(col)
            return row[i] if i is not None else None

        order_number = (get("order_number") or "").strip()
        if not order_number:
            skipped_rows += 1
            errors.append(f"L{total_rows+1}: order_number vide → ignorée")
            continue

        grouped[order_number].append({
            "order_number": order_number,
            "company_name": get("company_name"),
            "order_date": get("order_date"),
            "delivery_date": get("delivery_date"),
            "agent_raw": get("agent"),
            "sku": get("sku"),
            "qty": get("qty"),
            "unit_ht": get("unit_ht"),
            "code_client_raw": get("code_client"),
            "comment": get("comment") if has_comment else None,
        })

    created_orders = 0
    updated_orders = 0
    summaries = []

    for order_number, rows in grouped.items():
        sample = rows[0]

        # ----- agent -----
        agent_key = normalize_agent_key(sample["agent_raw"])
        agent = agent_index.get(agent_key)
        if not agent:
            errors.append(f"{order_number}: agent '{sample['agent_raw']}' introuvable → ignoré")
            skipped_rows += len(rows)
            continue

        # ----- client -----
        cc_key = normalize_code_client(sample["code_client_raw"])
        client = client_index.get(cc_key)
        if not client:
            errors.append(f"{order_number}: client '{sample['code_client_raw']}' introuvable → ignoré")
            skipped_rows += len(rows)
            continue

        # dates
        od = parse_date_safe(sample["order_date"])
        dd = parse_date_safe(sample["delivery_date"])

        # existing order
        res = await session.execute(
            select(Order)
            .options(selectinload(Order.items))
            .where(Order.labo_id == labo_id, Order.order_number == order_number)
        )
        order = res.scalar_one_or_none()

        if order:
            updated_orders += 1
            order.items.clear()
        else:
            order = Order(labo_id=labo_id, order_number=order_number)
            session.add(order)
            created_orders += 1

        # update header
        order.agent_id = agent.id
        order.client_id = client.id
        order.client_name = sample["company_name"] or client.company_name
        order.order_date = od
        order.delivery_date = dd
        order.currency = "EUR"
        order.comment = sample["comment"]  # colonne Excel

               # ---- totaux commande ----
        total_ht = Decimal("0")
        total_tva = Decimal("0")

        # items
        for r in rows:
            sku_raw = (r["sku"] or "").strip()
            if not sku_raw:
                errors.append(f"{order_number}: ligne sans SKU → ignorée")
                skipped_rows += 1
                continue

            sku_key = normalize_sku(sku_raw)
            product_id = None
            vat_rate = Decimal("0")

            prod_info = product_index.get(sku_key)
            if prod_info is not None:
                product_id, vat_rate = prod_info
            else:
                # On garde la ligne, mais sans product_id → pas de commission + TVA 0
                errors.append(
                    f"{order_number}: SKU '{sku_raw}' introuvable dans le catalogue labo {labo_id} "
                    f"→ product_id NULL (pas de commission / TVA à 0 sur cette ligne)"
                )

            qty = int(r["qty"] or 0)
            unit_ht = parse_decimal_safe(r["unit_ht"])
            line_ht = (unit_ht * qty).quantize(Decimal("0.01"))

            # TVA basée sur product.vat_rate (ex: 20.00)
            line_tva = (line_ht * vat_rate / Decimal("100")).quantize(Decimal("0.01"))

            total_ht += line_ht
            total_tva += line_tva

            item = OrderItem(
                product_id=product_id,
                sku=sku_raw,
                ean13=None,
                qty=qty,
                unit_ht=unit_ht,
                price_ht=unit_ht,   # 👈 PU HT (comme dans create_order)
                total_ht=line_ht,   # total ligne HT
                line_ht=line_ht,    # champ NOT NULL
            )
            order.items.append(item)

        # Totaux commande
        order.total_ht = total_ht.quantize(Decimal("0.01"))
        order.total_ttc = (total_ht + total_tva).quantize(Decimal("0.01"))


        summaries.append(
            ImportOrderSummary(
                order_number=order_number,
                client_id=client.id,
                agent_id=agent.id,
                total_ht=float(total_ht),
            )
        )

    await session.commit()

    return ImportResult(
        labo_id=labo_id,
        total_rows=total_rows,
        created_orders=created_orders,
        updated_orders=updated_orders,
        skipped_rows=skipped_rows,
        errors=errors,
        orders=summaries,
    )
