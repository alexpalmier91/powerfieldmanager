# app/routers/superuser_display_products_api.py
from __future__ import annotations

from typing import List, Optional, Any
import io

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    UploadFile,
    File,
    Form,
)
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from openpyxl import load_workbook

from app.db.session import get_async_session
from app.db.models import (
    DisplayProduct,
    RfidTagProductLink,
    DisplayOwnerClient,
    RfidTag,
)
from app.core.security import require_role

from app.schemas.display_products import (
    DisplayProductCreate,
    DisplayProductOut,
    RfidTagLinkCreate,
    RfidTagLinkOut,
    UnassignedEpcOut,
)

router = APIRouter(
    prefix="/api-zenhub/superuser/display-products",
    tags=["superuser-display-products"],
)

# =========================================================
#               EPC NON ASSIGNÉS (AVANT /{id})
# =========================================================

@router.get("/unassigned-epc", response_model=List[UnassignedEpcOut])
async def get_unassigned_epc(
    search: Optional[str] = Query(None, description="Filtre EPC par substring"),
    limit: int = Query(50, ge=1, le=500),
    session: AsyncSession = Depends(get_async_session),
    _: Any = Depends(require_role(["SUPERUSER", "SUPERADMIN"])),
):
    """
    EPC vus dans rfid_tag mais non encore liés à un Display Product
    """
    linked_subq = select(RfidTagProductLink.epc).subquery()

    stmt = (
        select(RfidTag.epc, RfidTag.last_seen_at)
        .where(~RfidTag.epc.in_(select(linked_subq.c.epc)))
        .order_by(RfidTag.last_seen_at.desc().nullslast())
        .limit(limit)
    )

    if search:
        stmt = stmt.where(RfidTag.epc.ilike(f"%{search}%"))

    res = await session.execute(stmt)
    rows = res.all()

    return [
        UnassignedEpcOut(epc=row.epc, last_seen_at=row.last_seen_at)
        for row in rows
    ]


# =========================================================
#               IMPORT EXCEL (AVANT /{id})
# =========================================================

@router.post("/import-excel")
async def import_display_products_excel(
    owner_client_id: int = Form(...),
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    _: Any = Depends(require_role(["SUPERUSER", "SUPERADMIN"])),
):
    """
    Import de Display Products depuis un fichier Excel (.xlsx)

    En-têtes attendues :
    - Colonne A : SKU
    - Colonne B : Nom du produit
    - Colonne C : EAN (optionnel)
    """
    owner = await session.get(DisplayOwnerClient, owner_client_id)
    if not owner:
        raise HTTPException(status_code=400, detail="Owner client introuvable")

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Merci de fournir un fichier Excel (.xlsx)")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    created = 0
    updated = 0
    skipped = 0

    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue

        sku = (row[0] or "").strip() if len(row) >= 1 and row[0] else ""
        name = (row[1] or "").strip() if len(row) >= 2 and row[1] else ""
        ean = (row[2] or "").strip() if len(row) >= 3 and row[2] else None

        if not sku:
            skipped += 1
            continue

        stmt = select(DisplayProduct).where(
            DisplayProduct.owner_client_id == owner_client_id,
            DisplayProduct.sku == sku,
        )
        res = await session.execute(stmt)
        existing = res.scalars().first()

        if existing:
            if name:
                existing.name = name
            existing.ean13 = ean
            updated += 1
        else:
            dp = DisplayProduct(
                owner_client_id=owner_client_id,
                sku=sku,
                name=name or sku,
                ean13=ean,
            )
            session.add(dp)
            created += 1

    await session.commit()

    return {
        "status": "ok",
        "owner_client_id": owner_client_id,
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }


# =========================================================
#               CRUD DISPLAY PRODUCTS
# =========================================================

@router.post("", response_model=DisplayProductOut)
async def create_display_product(
    payload: DisplayProductCreate,
    session: AsyncSession = Depends(get_async_session),
    _: Any = Depends(require_role(["SUPERUSER", "SUPERADMIN"])),
):
    owner = await session.get(DisplayOwnerClient, payload.owner_client_id)
    if not owner:
        raise HTTPException(status_code=400, detail="Owner client not found")

    stmt = select(DisplayProduct).where(
        DisplayProduct.owner_client_id == payload.owner_client_id,
        DisplayProduct.sku == payload.sku,
    )
    res = await session.execute(stmt)
    if res.scalars().first():
        raise HTTPException(
            status_code=400,
            detail="A display product with this SKU already exists for this owner",
        )

    dp = DisplayProduct(
        owner_client_id=payload.owner_client_id,
        sku=payload.sku,
        name=payload.name,
        description=payload.description,
    )
    session.add(dp)
    await session.commit()

    # ✅ re-fetch avec eager loading pour éviter MissingGreenlet pendant la sérialisation
    stmt = (
        select(DisplayProduct)
        .options(selectinload(DisplayProduct.owner_client))
        .where(DisplayProduct.id == dp.id)
    )
    res = await session.execute(stmt)
    dp2 = res.scalars().first()
    if not dp2:
        raise HTTPException(status_code=500, detail="Failed to load created product")
    return dp2


@router.get("", response_model=List[DisplayProductOut])
async def list_display_products(
    owner_client_id: Optional[int] = Query(None),
    q: Optional[str] = Query(None),
    session: AsyncSession = Depends(get_async_session),
    _: Any = Depends(require_role(["SUPERUSER", "SUPERADMIN"])),
):
    stmt = (
        select(DisplayProduct)
        .options(selectinload(DisplayProduct.owner_client))  # ✅ évite lazy-load Pydantic
        .order_by(DisplayProduct.created_at.desc())
    )

    if owner_client_id is not None:
        stmt = stmt.where(DisplayProduct.owner_client_id == owner_client_id)

    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            (DisplayProduct.sku.ilike(like))
            | (DisplayProduct.name.ilike(like))
        )

    res = await session.execute(stmt)
    return res.scalars().all()


@router.get("/{display_product_id}", response_model=DisplayProductOut)
async def get_display_product(
    display_product_id: int,
    session: AsyncSession = Depends(get_async_session),
    _: Any = Depends(require_role(["SUPERUSER", "SUPERADMIN"])),
):
    stmt = (
        select(DisplayProduct)
        .options(selectinload(DisplayProduct.owner_client))  # ✅ évite lazy-load Pydantic
        .where(DisplayProduct.id == display_product_id)
    )
    res = await session.execute(stmt)
    dp = res.scalars().first()

    if not dp:
        raise HTTPException(status_code=404, detail="Display product not found")
    return dp


@router.post("/{display_product_id}/link-epc", response_model=RfidTagLinkOut)
async def link_epc_to_display_product(
    display_product_id: int,
    payload: RfidTagLinkCreate,
    session: AsyncSession = Depends(get_async_session),
    _: Any = Depends(require_role(["SUPERUSER", "SUPERADMIN"])),
):
    dp = await session.get(DisplayProduct, display_product_id)
    if not dp:
        raise HTTPException(status_code=404, detail="Display product not found")

    epc = payload.epc.strip()
    if not epc:
        raise HTTPException(status_code=400, detail="EPC is required")

    stmt = select(RfidTagProductLink).where(RfidTagProductLink.epc == epc)
    res = await session.execute(stmt)
    if res.scalars().first():
        raise HTTPException(
            status_code=400,
            detail="This EPC is already linked to a display product",
        )

    link = RfidTagProductLink(
        epc=epc,
        display_product_id=display_product_id,
    )
    session.add(link)
    await session.commit()
    await session.refresh(link)
    return link
