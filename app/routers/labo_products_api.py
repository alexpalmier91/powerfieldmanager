from __future__ import annotations
from typing import List, Optional
import decimal
import os
from pathlib import Path
from datetime import date, timedelta  # 👈 ajouté

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import JSONResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from sqlalchemy import text, desc
from openpyxl import load_workbook
from io import BytesIO

from pydantic import BaseModel, conint, condecimal

import httpx  # <--- pour télécharger les images distantes

from app.db.session import get_async_session
from app.db.models import Labo, Product, PriceTier, User, UserRole
from app.core.security import get_current_subject


router = APIRouter(
    prefix="/api-zenhub/labo/products",
    tags=["labo-products"],
)

# Répertoire racine pour stocker les images produits labo
MEDIA_ROOT = Path(os.environ.get("MEDIA_ROOT", "media"))

# Base du projet (/srv/zenhub) à partir de ce fichier app/routers/labo_products_api.py
BASE_DIR = Path(__file__).resolve().parent.parent.parent

# Chemin absolu vers le modèle d'import produits
# => /srv/zenhub/app/static/labo/mode_labo_import_produits_example.xlsx
TEMPLATE_IMPORT_PATH = BASE_DIR / "app" / "static" / "labo" / "mode_labo_import_produits_example.xlsx"

LABO_DOC_TABLE = "labo_document"
LABO_DOC_ITEM_TABLE = "labo_document_item"
CLIENT_TABLE = "client"


# ======================
#   Dépendances
# ======================

async def get_current_labo(
    subject: str = Depends(get_current_subject),
    session: AsyncSession = Depends(get_async_session),
) -> Labo:
    """
    Récupère le labo courant à partir de l'email (subject du JWT).
    On charge l'objet User en BDD puis on vérifie son rôle / labo_id.
    """
    # Charger l'utilisateur actif correspondant à l'email
    res = await session.execute(
        select(User).where(User.email == subject, User.is_active == True)
    )
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=401, detail="Utilisateur introuvable ou inactif")

    if user.role not in (UserRole.LABO, UserRole.SUPERUSER):
        raise HTTPException(status_code=403, detail="Accès réservé au mode Labo")

    if not user.labo_id:
        raise HTTPException(status_code=400, detail="Aucun labo associé à l'utilisateur")

    labo = await session.get(Labo, user.labo_id)
    if not labo:
        raise HTTPException(status_code=404, detail="Labo introuvable")

    return labo


# ======================
#   Schemas Pydantic
# ======================

class ProductVariantOut(BaseModel):
    """Simple enveloppe pour coller au besoin 'Variantes' côté front."""
    ean13: Optional[str]
    price_ht: decimal.Decimal
    stock: int


class ProductOut(BaseModel):
    id: int
    sku: str
    name: str
    price_ht: decimal.Decimal
    stock: int
    ean13: Optional[str]
    is_active: bool
    variants: List[ProductVariantOut]
    # Taux de commission (en %) si besoin pour d'autres endpoints
    commission: Optional[decimal.Decimal] = None

    class Config:
        from_attributes = True


class ProductListResponse(BaseModel):
    items: List[ProductOut]
    total: int
    page: int
    page_size: int


class TierPriceIn(BaseModel):
    id: Optional[int] = None
    min_qty: conint(ge=1)
    price_ht: condecimal(max_digits=12, decimal_places=2)


class TierPriceOut(BaseModel):
    id: int
    min_qty: int
    price_ht: decimal.Decimal

    class Config:
        from_attributes = True


# 👉 Nouveaux schémas pour la commission

class ProductCommissionUpdate(BaseModel):
    # Commission en %, 0 à 100 avec 2 décimales
    commission: condecimal(max_digits=5, decimal_places=2, ge=0, le=100)


class ProductCommissionOut(BaseModel):
    id: int
    commission: decimal.Decimal

    class Config:
        from_attributes = True
        
        
# ======================
#   Schemas Stats produit
# ======================

class ProductHeader(BaseModel):
    id: int
    sku: str
    name: str
    ean13: Optional[str] = None
    labo_name: Optional[str] = None
    price_ht: Optional[decimal.Decimal] = None
    stock: Optional[int] = None


class GlobalStats(BaseModel):
    product: ProductHeader
    total_revenue_ht: decimal.Decimal
    total_qty: decimal.Decimal
    avg_price_ht: decimal.Decimal
    nb_clients: int
    first_sale_date: Optional[date]
    last_sale_date: Optional[date]


class MonthlyPoint(BaseModel):
    month: date  # premier jour du mois
    qty: decimal.Decimal
    revenue_ht: decimal.Decimal


class MonthlySalesResponse(BaseModel):
    points: list[MonthlyPoint]


class TopClientItem(BaseModel):
    client_id: int
    client_name: str
    total_qty: decimal.Decimal
    total_revenue_ht: decimal.Decimal
    last_purchase_date: Optional[date]


class TopClientsResponse(BaseModel):
    items: list[TopClientItem]


class SaleItem(BaseModel):
    date_document: date
    doc_type: str
    doc_number: str
    client_name: str
    qty: decimal.Decimal
    unit_price_ht: decimal.Decimal
    total_ht: decimal.Decimal


class SalesListResponse(BaseModel):
    items: list[SaleItem]
        

# ======================
#   Helpers Stats
# ======================

def _last_12_months_range() -> tuple[date, date]:
    """
    Fenêtre glissante ~12 derniers mois (approche simple).
    """
    today = date.today()
    start = (today.replace(day=1) - timedelta(days=365))
    end = today + timedelta(days=1)
    return start, end


# ======================
#   Listing produits
# ======================

@router.get("")
async def list_products(
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: Optional[str] = Query("sku"),   # tri par défaut SKU
    sort_dir: Optional[str] = Query("asc"),  # asc / desc
):
    print(">>> list_products() labo_products_api.py appelé, page =", page)

    stmt = select(Product).where(Product.labo_id == labo.id)

    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(
            or_(
                Product.sku.ilike(pattern),
                Product.name.ilike(pattern),
                Product.ean13.ilike(pattern),
            )
        )

    # =============== TRI SERVEUR ===============
    sort_col = Product.id
    sort_by = (sort_by or "").lower()
    if sort_by == "sku":
        sort_col = Product.sku
    elif sort_by == "name":
        sort_col = Product.name
    elif sort_by == "stock":
        sort_col = Product.stock

    if (sort_dir or "").lower() == "desc":
        sort_col = sort_col.desc()
    else:
        sort_col = sort_col.asc()

    stmt = (
        stmt.order_by(sort_col, Product.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    # ===========================================

    result = await session.execute(stmt)
    products = result.scalars().all()

    # ===== Récupération des paliers de prix pour les produits chargés =====
    product_ids = [p.id for p in products]
    tiers_map: dict[int, list[dict]] = {}
    tiers_count_map: dict[int, int] = {}

    if product_ids:
        tiers_stmt = (
            select(
                PriceTier.product_id,
                PriceTier.qty_min,
                PriceTier.price_ht,
            )
            .where(PriceTier.product_id.in_(product_ids))
            .order_by(PriceTier.product_id, PriceTier.qty_min)
        )
        tiers_res = await session.execute(tiers_stmt)
        for pid, qty_min, price_ht in tiers_res.all():
            tiers_map.setdefault(pid, []).append(
                {
                    "min_qty": int(qty_min),
                    "price_ht": float(price_ht or 0),
                }
            )

        for pid, tiers in tiers_map.items():
            tiers_count_map[pid] = len(tiers)
    # =====================================================================

    items: list[dict] = []
    for p in products:
        img_url = p.image_url or f"/media/labo_products/{p.labo_id}/{p.sku}.jpg"
        product_tiers = tiers_map.get(p.id, [])
        tiers_count = tiers_count_map.get(p.id, 0)

        items.append(
            {
                "id": p.id,
                "sku": p.sku,
                "name": p.name,
                "price_ht": float(p.price_ht or 0),
                "stock": p.stock or 0,
                "ean13": p.ean13,
                "image_url": img_url,
                "is_active": bool(p.is_active),
                "commission": float(p.commission or 0),
                "tiers": product_tiers,          # 👈 tous les paliers pour affichage
                "tiers_count": tiers_count,      # 👈 compteur pour info si besoin
                "variants": [
                    {
                        "ean13": p.ean13,
                        "price_ht": float(p.price_ht or 0),
                        "stock": p.stock or 0,
                    }
                ],
            }
        )

    return items
    

# ======================
#   STATS PRODUIT
# ======================

@router.get("/{product_id}/stats/global", response_model=GlobalStats)
async def product_global_stats(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    # Récup produit
    stmt_prod = select(Product).where(
        Product.id == product_id,
        Product.labo_id == labo.id,
    )
    product = (await session.execute(stmt_prod)).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")

    # Agrégations sur les factures FA (SQL brut)
    sql = text(f"""
        SELECT
          COALESCE(SUM(i.qty), 0) AS total_qty,
          COALESCE(SUM(i.total_ht), 0) AS total_revenue_ht,
          COUNT(DISTINCT d.client_id) AS nb_clients,
          MIN(d.order_date) AS first_sale_date,
          MAX(d.order_date) AS last_sale_date,
          COALESCE(SUM(i.total_ht), 0) / NULLIF(COALESCE(SUM(i.qty), 0), 0) AS avg_price_ht
        FROM {LABO_DOC_ITEM_TABLE} i
        JOIN {LABO_DOC_TABLE} d ON d.id = i.document_id
        WHERE i.product_id = :product_id
          AND d.labo_id = :labo_id
          AND d.type = 'FA'
    """)

    res = await session.execute(sql, {"product_id": product_id, "labo_id": labo.id})
    row = res.mappings().one()

    header = ProductHeader(
        id=product.id,
        sku=product.sku,
        name=product.name,
        ean13=product.ean13,
        labo_name=labo.name,
        price_ht=product.price_ht or decimal.Decimal("0.00"),
        stock=product.stock or 0,
    )

    return GlobalStats(
        product=header,
        total_revenue_ht=row["total_revenue_ht"] or decimal.Decimal("0.00"),
        total_qty=row["total_qty"] or decimal.Decimal("0.00"),
        avg_price_ht=row["avg_price_ht"] or decimal.Decimal("0.00"),
        nb_clients=int(row["nb_clients"] or 0),
        first_sale_date=row["first_sale_date"],
        last_sale_date=row["last_sale_date"],
    )




@router.get("/{product_id}/stats/monthly-sales", response_model=MonthlySalesResponse)
async def product_monthly_sales(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    start, end = _last_12_months_range()

    sql = text(f"""
        SELECT
          DATE_TRUNC('month', d.order_date) AS month,
          COALESCE(SUM(i.qty), 0) AS qty,
          COALESCE(SUM(i.total_ht), 0) AS revenue_ht
        FROM {LABO_DOC_ITEM_TABLE} i
        JOIN {LABO_DOC_TABLE} d ON d.id = i.document_id
        WHERE i.product_id = :product_id
          AND d.labo_id = :labo_id
          AND d.type = 'FA'
          AND d.order_date >= :start
          AND d.order_date < :end
        GROUP BY DATE_TRUNC('month', d.order_date)
        ORDER BY DATE_TRUNC('month', d.order_date)
    """)



    res = await session.execute(
        sql,
        {"product_id": product_id, "labo_id": labo.id, "start": start, "end": end},
    )
    rows = res.mappings().all()

    points: list[MonthlyPoint] = []
    for row in rows:
        month_dt = row["month"]
        # month_dt est un datetime → on le force au 1er du mois
        month_date = month_dt.date().replace(day=1)
        points.append(
            MonthlyPoint(
                month=month_date,
                qty=row["qty"] or decimal.Decimal("0.00"),
                revenue_ht=row["revenue_ht"] or decimal.Decimal("0.00"),
            )
        )

    return MonthlySalesResponse(points=points)



@router.get("/{product_id}/stats/monthly-revenue", response_model=MonthlySalesResponse)
async def product_monthly_revenue(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    # même payload, le front exploitera revenue_ht
    return await product_monthly_sales(product_id, labo, session)



@router.get("/{product_id}/stats/top-clients", response_model=TopClientsResponse)
async def product_top_clients(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
    limit: int = 10,
):
    sql = text(f"""
        SELECT
          c.id AS client_id,
          c.company_name AS client_name,
          COALESCE(SUM(i.qty), 0) AS total_qty,
          COALESCE(SUM(i.total_ht), 0) AS total_revenue_ht,
          MAX(d.order_date) AS last_purchase_date
        FROM {LABO_DOC_ITEM_TABLE} i
        JOIN {LABO_DOC_TABLE} d ON d.id = i.document_id
        JOIN {CLIENT_TABLE} c ON c.id = d.client_id
        WHERE i.product_id = :product_id
          AND d.labo_id = :labo_id
          AND d.type = 'FA'
        GROUP BY c.id, c.company_name
        ORDER BY total_revenue_ht DESC
        LIMIT :limit
    """)



    res = await session.execute(
        sql,
        {"product_id": product_id, "labo_id": labo.id, "limit": limit},
    )
    rows = res.mappings().all()

    items: list[TopClientItem] = []
    for row in rows:
        items.append(
            TopClientItem(
                client_id=row["client_id"],
                client_name=row["client_name"],
                total_qty=row["total_qty"] or decimal.Decimal("0.00"),
                total_revenue_ht=row["total_revenue_ht"] or decimal.Decimal("0.00"),
                last_purchase_date=row["last_purchase_date"],
            )
        )

    return TopClientsResponse(items=items)




@router.get("/{product_id}/stats/sales-list", response_model=SalesListResponse)
async def product_sales_list(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
    limit: int = 500,
):
    sql = text(f"""
        SELECT
          d.order_date AS date_document,
          d.type AS doc_type,
          d.order_number AS doc_number,
          c.company_name AS client_name,
          i.qty,
          i.unit_ht AS unit_price_ht,
          i.total_ht
        FROM {LABO_DOC_ITEM_TABLE} i
        JOIN {LABO_DOC_TABLE} d ON d.id = i.document_id
        JOIN {CLIENT_TABLE} c ON c.id = d.client_id
        WHERE i.product_id = :product_id
          AND d.labo_id = :labo_id
          AND d.type IN ('FA', 'BL', 'BC')
        ORDER BY d.order_date DESC, d.order_number DESC
        LIMIT :limit
    """)


    res = await session.execute(
        sql,
        {"product_id": product_id, "labo_id": labo.id, "limit": limit},
    )
    rows = res.mappings().all()  # 👈 important: .mappings()

    items: list[SaleItem] = []
    for row in rows:
        items.append(
            SaleItem(
                date_document=row["date_document"],
                doc_type=row["doc_type"],
                doc_number=row["doc_number"],
                client_name=row["client_name"],
                qty=row["qty"] or decimal.Decimal("0.00"),
                unit_price_ht=row["unit_price_ht"] or decimal.Decimal("0.00"),
                total_ht=row["total_ht"] or decimal.Decimal("0.00"),
            )
        )

    return SalesListResponse(items=items)









@router.get("/debug/{product_id}")
async def debug_product(
    product_id: int,
    session: AsyncSession = Depends(get_async_session),
):
    # On lit DIRECTEMENT la table product avec du SQL brut
    res = await session.execute(
        text("SELECT id, sku, image_url FROM product WHERE id = :pid"),
        {"pid": product_id},
    )
    row = res.mappings().first()
    if not row:
        return {"found": False}

    return {
        "found": True,
        "id": row["id"],
        "sku": row["sku"],
        "image_url": row["image_url"],
    }


# ======================
#   Toggle is_active
# ======================

@router.post("/{product_id}/toggle-active")
async def toggle_active(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    stmt = select(Product).where(
        Product.id == product_id,
        Product.labo_id == labo.id,
    )
    product = (await session.execute(stmt)).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")

    product.is_active = not product.is_active
    await session.commit()
    await session.refresh(product)

    return {"id": product.id, "is_active": product.is_active}


# ======================
#   PATCH commission
# ======================

@router.patch("/{product_id}/commission", response_model=ProductCommissionOut)
async def update_commission(
    product_id: int,
    payload: ProductCommissionUpdate,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    """
    Mise à jour du taux de commission pour un produit du labo connecté.

    - Vérifie que le produit appartient au labo courant.
    - Enregistre la nouvelle commission en BDD.
    - Retourne la valeur mise à jour.
    """
    stmt = select(Product).where(
        Product.id == product_id,
        Product.labo_id == labo.id,
    )
    product = (await session.execute(stmt)).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")

    # Mise à jour
    product.commission = payload.commission

    await session.commit()
    await session.refresh(product)

    return ProductCommissionOut(
        id=product.id,
        commission=product.commission or decimal.Decimal("0.00"),
    )


# ======================
#   Tiers price
# ======================

@router.get("/{product_id}/tiers", response_model=list[TierPriceOut])
async def get_tiers(
    product_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    # Vérifier que le produit est dans le labo
    prod_stmt = select(Product).where(
        Product.id == product_id,
        Product.labo_id == labo.id,
    )
    product = (await session.execute(prod_stmt)).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")

    stmt = (
        select(PriceTier)
        .where(PriceTier.product_id == product_id)
        .order_by(PriceTier.qty_min)
    )
    tiers = (await session.execute(stmt)).scalars().all()

    return [
        TierPriceOut(
            id=t.id,
            min_qty=t.qty_min,
            price_ht=t.price_ht,
        )
        for t in tiers
    ]


@router.post("/{product_id}/tiers", response_model=TierPriceOut)
async def upsert_tier(
    product_id: int,
    payload: TierPriceIn,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    # Vérifier que le produit est dans le labo
    prod_stmt = select(Product).where(
        Product.id == product_id,
        Product.labo_id == labo.id,
    )
    product = (await session.execute(prod_stmt)).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Produit introuvable")

    if payload.id:
        stmt = select(PriceTier).where(
            PriceTier.id == payload.id,
            PriceTier.product_id == product_id,
        )
        tier = (await session.execute(stmt)).scalar_one_or_none()
        if not tier:
            raise HTTPException(status_code=404, detail="Palier introuvable")
        tier.qty_min = payload.min_qty
        tier.price_ht = payload.price_ht
    else:
        # upsert par (product_id, qty_min)
        stmt = select(PriceTier).where(
            PriceTier.product_id == product_id,
            PriceTier.qty_min == payload.min_qty,
        )
        tier = (await session.execute(stmt)).scalar_one_or_none()
        if tier:
            tier.price_ht = payload.price_ht
        else:
            tier = PriceTier(
                product_id=product.id,
                qty_min=payload.min_qty,
                price_ht=payload.price_ht,
            )
            session.add(tier)

    await session.commit()
    await session.refresh(tier)
    return TierPriceOut(
        id=tier.id,
        min_qty=tier.qty_min,
        price_ht=tier.price_ht,
    )


@router.delete("/{product_id}/tiers/{tier_id}")
async def delete_tier(
    product_id: int,
    tier_id: int,
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    # Vérifier qu'il appartient bien au labo via le produit
    stmt = (
        select(PriceTier)
        .join(Product, PriceTier.product_id == Product.id)
        .where(
            PriceTier.id == tier_id,
            PriceTier.product_id == product_id,
            Product.labo_id == labo.id,
        )
    )
    tier = (await session.execute(stmt)).scalar_one_or_none()
    if not tier:
        raise HTTPException(status_code=404, detail="Palier introuvable")

    await session.delete(tier)
    await session.commit()
    return {"ok": True}


# ======================
#   Téléchargement modèle import
# ======================

@router.get("/import-template")
async def download_import_template():
    """
    Télécharge le fichier Excel modèle pour l'import produits.
    URL : /api-zenhub/labo/products/import-template
    """
    if not TEMPLATE_IMPORT_PATH.exists():
        raise HTTPException(status_code=404, detail="Fichier modèle introuvable")

    return FileResponse(
        path=str(TEMPLATE_IMPORT_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="mode_labo_import_produits_example.xlsx",
    )


# ======================
#   Import produits (+ image_url distante)
# ======================

@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    labo: Labo = Depends(get_current_labo),
    session: AsyncSession = Depends(get_async_session),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Fichier .xlsx attendu")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    # Lecture entêtes
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, col_name in enumerate(row):
            if col_name is None:
                continue
            headers[str(col_name).strip()] = idx
        break

    # Colonnes obligatoires (ajout de vat_rate)
    required_cols = ["sku", "name", "price_ht", "stock", "image_url", "vat_rate"]
    for col in required_cols:
        if col not in headers:
            return JSONResponse(
                status_code=400,
                content={
                    "created": 0,
                    "updated": 0,
                    "errors": [
                        {
                            "row": 1,
                            "sku": "",
                            "message": f"Colonne obligatoire manquante : {col}",
                        }
                    ],
                },
            )

    created = 0
    updated = 0
    errors: list[dict] = []

    def get_val(row, col_name, default=None):
        idx = headers.get(col_name)
        if idx is None:
            return default
        return row[idx]

    # Répertoire pour les images de ce labo
    labo_image_dir = MEDIA_ROOT / "labo_products" / str(labo.id)
    labo_image_dir.mkdir(parents=True, exist_ok=True)

    # On utilise un seul client HTTP pour tout le fichier
    async with httpx.AsyncClient(timeout=20.0) as client:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            sku = get_val(row, "sku")
            if not sku or str(sku).strip() == "":
                errors.append({"row": row_idx, "sku": "", "message": "SKU manquant"})
                continue
            sku = str(sku).strip()

            name = get_val(row, "name") or ""
            description = get_val(row, "description") or ""
            price_raw = get_val(row, "price_ht")
            stock_raw = get_val(row, "stock")
            ean13_raw = get_val(row, "ean13")
            is_active_raw = get_val(row, "is_active")
            tier_min_qty_raw = get_val(row, "tier_min_qty")
            tier_price_ht_raw = get_val(row, "tier_price_ht")
            image_url_raw = get_val(row, "image_url")
            vat_rate_raw = get_val(row, "vat_rate")

            # price_ht
            try:
                price_ht = decimal.Decimal(str(price_raw).replace(",", ".")) if price_raw is not None else None
            except Exception:
                errors.append({"row": row_idx, "sku": sku, "message": "price_ht invalide"})
                continue

            # stock
            try:
                stock = int(stock_raw) if stock_raw is not None else 0
                if stock < 0:
                    raise ValueError()
            except Exception:
                errors.append({"row": row_idx, "sku": sku, "message": "stock invalide"})
                continue

            # vat_rate (0–100, par ex. 20, 5.5…)
            try:
                if vat_rate_raw in (None, ""):
                    # si vide, on met le défaut de la BDD (20.00)
                    vat_rate = decimal.Decimal("20.00")
                else:
                    vat_rate = decimal.Decimal(str(vat_rate_raw).replace(",", "."))
                if vat_rate < 0 or vat_rate > 100:
                    raise ValueError()
            except Exception:
                errors.append(
                    {"row": row_idx, "sku": sku, "message": "vat_rate invalide (0–100)"}
                )
                continue

            # ean13
            ean13 = None
            if ean13_raw not in (None, ""):
                ean13 = str(ean13_raw).strip()
                if not (len(ean13) == 13 and ean13.isdigit()):
                    errors.append(
                        {"row": row_idx, "sku": sku, "message": "EAN13 invalide (13 chiffres)"}
                    )
                    continue

            # is_active
            is_active = True
            if is_active_raw not in (None, ""):
                try:
                    is_active = bool(int(is_active_raw))
                except Exception:
                    errors.append(
                        {"row": row_idx, "sku": sku, "message": "is_active invalide (0/1)"}
                    )
                    continue

            # Produit existant ?
            stmt_prod = select(Product).where(
                Product.labo_id == labo.id,
                Product.sku == sku,
            )
            product = (await session.execute(stmt_prod)).scalar_one_or_none()

            if product:
                product.name = str(name)
                product.description = str(description)
                product.price_ht = price_ht
                product.stock = stock
                product.ean13 = ean13
                product.is_active = is_active
                product.vat_rate = vat_rate
                updated += 1
            else:
                product = Product(
                    labo_id=labo.id,
                    sku=sku,
                    name=str(name),
                    description=str(description),
                    price_ht=price_ht,
                    stock=stock,
                    ean13=ean13,
                    is_active=is_active,
                    vat_rate=vat_rate,
                )
                session.add(product)
                await session.flush()
                created += 1

            # Palier tiers éventuel
            if tier_min_qty_raw not in (None, "") and tier_price_ht_raw not in (None, ""):
                try:
                    tier_min_qty = int(tier_min_qty_raw)
                    if tier_min_qty <= 0:
                        raise ValueError()
                    tier_price_ht = decimal.Decimal(str(tier_price_ht_raw).replace(",", "."))
                except Exception:
                    errors.append(
                        {
                            "row": row_idx,
                            "sku": sku,
                            "message": "tier_min_qty/tier_price_ht invalides",
                        }
                    )
                else:
                    stmt_tier = select(PriceTier).where(
                        PriceTier.product_id == product.id,
                        PriceTier.qty_min == tier_min_qty,
                    )
                    tier = (await session.execute(stmt_tier)).scalar_one_or_none()
                    if tier:
                        tier.price_ht = tier_price_ht
                    else:
                        tier = PriceTier(
                            product_id=product.id,
                            qty_min=tier_min_qty,
                            price_ht=tier_price_ht,
                        )
                        session.add(tier)

            # Téléchargement de l'image distante si image_url renseignée
            remote_image_url = None
            if image_url_raw not in (None, ""):
                remote_image_url = str(image_url_raw).strip()

            if remote_image_url:
                try:
                    resp = await client.get(remote_image_url)
                    if resp.status_code != 200:
                        errors.append(
                            {
                                "row": row_idx,
                                "sku": sku,
                                "message": f"Échec téléchargement image ({resp.status_code})",
                            }
                        )
                    else:
                        # Déterminer l'extension du fichier
                        ext = ".jpg"
                        url_path = remote_image_url.split("?", 1)[0]
                        if "." in url_path:
                            ext_candidate = url_path.rsplit(".", 1)[-1].lower()
                            if ext_candidate in ("jpg", "jpeg", "png", "webp"):
                                ext = "." + ext_candidate

                        filename = f"{sku}{ext}"
                        dest_path = labo_image_dir / filename
                        with open(dest_path, "wb") as f:
                            f.write(resp.content)

                        # URL stockée en BDD (adapter si ton prefix static diffère)
                        product.image_url = f"/media/labo_products/{labo.id}/{filename}"

                except Exception as e:
                    errors.append(
                        {
                            "row": row_idx,
                            "sku": sku,
                            "message": f"Erreur téléchargement image: {e}",
                        }
                    )

    await session.commit()

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
    }
