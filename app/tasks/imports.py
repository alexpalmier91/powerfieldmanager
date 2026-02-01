import os
import csv
import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import create_engine, select, func, text, inspect
from sqlalchemy.orm import sessionmaker

from app.tasks.celery_app import celery
from app.db.models import Product, ImportJob, Agent, LaboClient  # 👈 ajout LaboClient
# Modèles optionnels (si non exposés)
try:
    from app.db.models import Order, OrderItem, Customer
except Exception:
    Order = None
    OrderItem = None
    Customer = None

# -------------------------------------------------------------------
# Connexions synchrones pour Celery (pas d'async dans un worker)
# -------------------------------------------------------------------
DATABASE_URL_SYNC = os.getenv("DATABASE_URL", "").replace("+asyncpg", "")
engine = create_engine(DATABASE_URL_SYNC, pool_pre_ping=True, future=True)
Session = sessionmaker(bind=engine, expire_on_commit=False)
inspector = inspect(engine)

LABO_A_ID_DEFAULT = 1  # par défaut pour Labo A

# -------------------------------------------------------------------
# Utils
# -------------------------------------------------------------------
def _norm(s: str) -> str:
    import unicodedata, re
    s = (s or "").strip()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")

def _parse_date_any(x) -> Optional[datetime.date]:
    if x is None or str(x).strip() == "":
        return None
    try:
        return pd.to_datetime(x, dayfirst=True, errors="coerce").date()
    except Exception:
        return None

def _safe_float(x) -> Optional[float]:
    if x is None or str(x).strip() == "":
        return None
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None

def _get_table_columns(table_name: str) -> set:
    try:
        cols = inspector.get_columns(table_name)
        return set(c["name"] for c in cols)
    except Exception:
        return set()

def _only_existing_fields(data: Dict[str, Any], table_name: str) -> Dict[str, Any]:
    cols = _get_table_columns(table_name)
    return {k: v for k, v in data.items() if k in cols}

def _save_job_progress(sess, job, inserted, updated, errors_list):
    job.inserted = inserted
    job.updated = updated
    job.errors = errors_list
    sess.commit()

# -------------------------------------------------------------------
# LECTURE FICHIERS (CSV / XLSX)
# -------------------------------------------------------------------
def _read_any(filepath: str) -> pd.DataFrame:
    ext = os.path.splitext(filepath.lower())[1]
    if ext == ".csv":
        return pd.read_csv(filepath, sep=None, engine="python")
    if ext in (".xlsx", ".xls"):
        return pd.read_excel(filepath)
    raise ValueError("Type de fichier non supporté (.csv / .xlsx / .xls)")

# -------------------------------------------------------------------
# Résolution des AGENTS (email prioritaire, sinon “Nom” OU “Prénom Nom”)
# -------------------------------------------------------------------
def _normalize_whitespace(s: Optional[str]) -> str:
    return " ".join((s or "").strip().split())

def _split_fullname(name: str) -> Tuple[Optional[str], Optional[str]]:
    name = _normalize_whitespace(name)
    if not name:
        return None, None
    parts = name.split(" ")
    if len(parts) == 1:
        return None, parts[0]  # NOM seul
    return parts[0], " ".join(parts[1:])  # (first, last)

def _find_agent_by_last_or_fullname(sess, agent_name: str) -> Optional[Agent]:
    if not agent_name:
        return None
    first, last = _split_fullname(agent_name)

    # Cas 1 : NOM seul
    if last and not first:
        q = select(Agent).where(func.lower(Agent.lastname) == func.lower(last)).limit(1)
        return sess.execute(q).scalar_one_or_none()

    # Cas 2 : Prénom + Nom
    if first and last:
        q = select(Agent).where(
            func.lower(Agent.firstname) == func.lower(first),
            func.lower(Agent.lastname) == func.lower(last),
        ).limit(1)
        a = sess.execute(q).scalar_one_or_none()
        if a:
            return a

    # Fallback : comparer tout sur last_name
    q = select(Agent).where(func.lower(Agent.lastname) == func.lower(last or agent_name)).limit(1)
    return sess.execute(q).scalar_one_or_none()

def _resolve_agent(sess, agent_email: Optional[str], agent_name: Optional[str]) -> Optional[Agent]:
    if agent_email:
        a = sess.execute(
            select(Agent).where(func.lower(Agent.email) == func.lower(agent_email))  # 👈 insensible casse
        ).scalar_one_or_none()
        if a:
            return a
    return _find_agent_by_last_or_fullname(sess, agent_name or "")

def _ensure_agent_labo(sess, agent_id: int, labo_id: int):
    """Lie l'agent au labo via labo_agent (idempotent)."""
    try:
        sess.execute(
            text("""
                INSERT INTO labo_agent(labo_id, agent_id)
                VALUES (:labo_id, :agent_id)
                ON CONFLICT DO NOTHING
            """),
            {"labo_id": labo_id, "agent_id": agent_id}
        )
        sess.flush()
    except Exception:
        sess.rollback()

def _ensure_labo_client(sess, labo_id: int, client_obj, code_client: Optional[str]):
    """Crée (si besoin) le lien labo ↔ client avec le code client propre au labo."""
    if not client_obj or not code_client:
        return
    try:
        sess.execute(
            text("""
                INSERT INTO labo_client (labo_id, client_id, code_client)
                VALUES (:labo_id, :client_id, :code_client)
                ON CONFLICT (labo_id, code_client) DO NOTHING
            """),
            {"labo_id": labo_id, "client_id": client_obj.id, "code_client": code_client}
        )
        sess.flush()
    except Exception:
        sess.rollback()

# -------------------------------------------------------------------
# CLIENT (facultatif : on détecte les colonnes dispo)
# -------------------------------------------------------------------
def _find_or_create_customer(sess, labo_id: int, name: Optional[str], code_client: Optional[str] = None):
    if Customer is None:
        return None

    name = _normalize_whitespace(name)
    code_client = _normalize_whitespace(code_client)

    customer_cols = _get_table_columns("customer")

    # match prioritaire par code client si la colonne existe
    if code_client and "code_client" in customer_cols:
        c = sess.execute(select(Customer).where(Customer.code_client == code_client)).scalar_one_or_none()
        if c:
            return c

    # match par nom si la colonne 'name' existe
    if name and "name" in customer_cols:
        c = sess.execute(select(Customer).where(func.lower(Customer.name) == func.lower(name))).scalar_one_or_none()
        if c:
            return c

    # création minimale si 'name' existe
    if "name" in customer_cols:
        c = Customer(name=name or "Client inconnu")
        if "labo_id" in customer_cols:
            setattr(c, "labo_id", labo_id)
        sess.add(c); sess.flush()
        return c

    return None

# -------------------------------------------------------------------
# === EXISTANT : Import PRODUITS (ajusté aux champs ImportJob) ======
# -------------------------------------------------------------------
def _normalize_row_from_raw_product(raw: Dict[str, Any]) -> Dict[str, Any]:
    ALIAS = {
        "reference": "sku", "référence": "sku", "ref": "sku",
        "article": "name", "nom": "name", "titre": "name",
        "desc": "description", "description": "description",
        "ean": "ean13", "ean13": "ean13",
        "prix_de_vente_ht": "price_ht", "prix_ht": "price_ht", "prix": "price_ht",
        "quantite": "stock", "quantité": "stock", "qty": "stock", "stock": "stock",
        "product_cover_image_url": "image_url", "image_url": "image_url", "image": "image_url",
    }

    normed: Dict[str, Any] = {}
    for k, v in raw.items():
        nk = _norm(k)
        normed[nk] = v

    out: Dict[str, Any] = {}
    for nk, v in normed.items():
        key = ALIAS.get(nk, nk)
        out[key] = v

    sku = str(out.get("sku") or "").strip()
    if not sku:
        return {}

    name = str(out.get("name") or sku).strip()
    description = (str(out.get("description")) if out.get("description") is not None else "").strip() or None
    image_url = (str(out.get("image_url")) if out.get("image_url") is not None else "").strip() or None

    ean_raw = out.get("ean13")
    ean13 = str(ean_raw).strip() if ean_raw is not None and str(ean_raw).strip() != "" else None

    price_raw = out.get("price_ht")
    price_ht = _safe_float(price_raw)

    stock_raw = out.get("stock")
    stock = None
    if stock_raw is not None and str(stock_raw).strip() != "":
        try:
            stock = int(float(stock_raw))
        except Exception:
            stock = None

    return {
        "sku": sku,
        "name": name,
        "description": description,
        "image_url": image_url,
        "ean13": ean13,
        "price_ht": price_ht,
        "stock": stock,
    }

@celery.task(name="import_products")
def import_products(filepath: str, labo_id: int | None = None):
    s = Session()
    task_id = import_products.request.id
    logger.info(f"[import_products] start task_id={task_id} file={filepath} labo_id={labo_id}")

    try:
        job = s.execute(select(ImportJob).where(ImportJob.task_id == task_id)).scalar_one_or_none()
        if not job:
            job = ImportJob(task_id=task_id, filename=os.path.basename(filepath))
            s.add(job); s.flush()
        else:
            job.task_id = task_id
        job.status = "STARTED"
        s.commit()

        rows_raw: List[Dict[str, Any]] = []
        lower = filepath.lower()
        if lower.endswith(".csv"):
            with open(filepath, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                rows_raw = list(reader)
        elif lower.endswith(".xlsx"):
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(values_only=True))]
            for row in ws.iter_rows(values_only=True, min_row=2):
                rows_raw.append(dict(zip(headers, row)))
        else:
            raise ValueError("Type de fichier non supporté (.csv ou .xlsx requis)")

        total = len(rows_raw)
        job.total_rows = total
        job.inserted = 0
        job.updated = 0
        job.errors = []
        s.commit()

        inserted = updated = 0
        errors: List[Dict[str, Any]] = []
        BATCH = 50

        for i, raw in enumerate(rows_raw, start=1):
            try:
                row = _normalize_row_from_raw_product(raw)
                if not row:
                    continue

                sku = row["sku"]
                name = row["name"]
                desc = row["description"]
                ean13 = row["ean13"]
                price_ht = row["price_ht"]
                stock = row["stock"]
                image_url = row["image_url"]

                prod = s.execute(
                    select(Product).where(Product.labo_id == labo_id, Product.sku == sku)
                ).scalar_one_or_none()

                if not prod:
                    prod = Product(
                        labo_id=labo_id, sku=sku, name=name, description=desc,
                        image_url=image_url, ean13=ean13 if ean13 else None,
                        price_ht=price_ht if price_ht is not None else 0,
                        stock=stock if stock is not None else 0,
                    )
                    s.add(prod); s.flush()
                    inserted += 1
                else:
                    prod.name = name
                    prod.description = desc
                    prod.image_url = image_url
                    if ean13: prod.ean13 = ean13
                    if price_ht is not None: prod.price_ht = price_ht
                    if stock is not None: prod.stock = stock
                    updated += 1

                if i % BATCH == 0:
                    s.flush()
                    _save_job_progress(s, job, inserted, updated, errors)
                    logger.info(f"[import_products] progress {i}/{total} (ins={inserted}, upd={updated}, err={len(errors)})")

            except Exception as e:
                s.rollback()
                errors.append({"row": i, "error": str(e)})
                job.errors = errors; s.commit()

        s.flush()
        job.inserted = inserted
        job.updated = updated
        job.errors = errors
        job.status = "SUCCESS" if not errors else "FAILURE"
        job.finished_at = datetime.datetime.utcnow()
        s.commit()

        logger.info(f"[import_products] done status={job.status} ins={inserted} upd={updated} total={total} err={len(errors)}")
        return {"status": job.status, "inserted": inserted, "updated": updated, "total": total, "errors": len(errors)}

    except Exception as e:
        logger.exception(f"[import_products] error {e}")
        s.rollback()
        if 'job' in locals() and job:
            job.status = "FAILURE"
            job.errors = [{"error": str(e)}]
            job.finished_at = datetime.datetime.utcnow()
            s.commit()
        raise
    finally:
        s.close()
        try:
            os.remove(filepath)
        except Exception:
            pass

# -------------------------------------------------------------------
# Import AGENTS
# -------------------------------------------------------------------
@celery.task(name="task_import_agents")
def task_import_agents(job_id: int, tmp_path: str, labo_id: int = LABO_A_ID_DEFAULT):
    s = Session()
    try:
        job = s.get(ImportJob, job_id)
        job.task_id = task_import_agents.request.id
        job.status = "STARTED"; s.commit()

        df = _read_any(tmp_path)
        cols = { _norm(c): c for c in df.columns }

        def pick(*names):
            for n in names:
                if _norm(n) in cols: return cols[_norm(n)]
            return None

        c_first = pick("first_name","firstname","prenom","prénom")
        c_last  = pick("last_name","lastname","nom")
        c_email = pick("email","mail")
        c_phone = pick("phone","telephone","téléphone")

        if not (c_first and c_last and c_email):
            raise RuntimeError("Colonnes minimales manquantes (first_name, last_name, email).")

        df = df.rename(columns={c_first:"first_name", c_last:"last_name", c_email:"email"})
        if c_phone: df = df.rename(columns={c_phone:"phone"})

        job.total_rows = int(len(df)); job.inserted = 0; job.updated = 0; job.errors = []; s.commit()
        inserted = updated = 0
        errors: List[Dict[str, Any]] = []

        for i, r in df.iterrows():
            try:
                email = str(r.get("email") or "").strip().lower()
                if not email:
                    continue
                first = _normalize_whitespace(r.get("first_name"))
                last  = _normalize_whitespace(r.get("last_name"))
                phone = str(r.get("phone")).strip() if "phone" in df.columns and pd.notna(r.get("phone")) else None

                ex = s.execute(select(Agent).where(Agent.email == email)).scalar_one_or_none()
                if ex:
                    ex.firstname = first or ex.firstname
                    ex.lastname  = last  or ex.lastname
                    ex.phone     = phone or ex.phone
                    updated += 1
                    agent_obj = ex
                else:
                    agent_obj = Agent(email=email, firstname=first, lastname=last, phone=phone)
                    s.add(agent_obj); s.flush()
                    inserted += 1

                _ensure_agent_labo(s, agent_obj.id, labo_id)
                if (i + 1) % 100 == 0:
                    _save_job_progress(s, job, inserted, updated, errors)

            except Exception as e:
                s.rollback()
                errors.append({"row": int(i)+1, "error": str(e)})
                job.errors = errors; s.commit()

        job.status = "SUCCESS" if not errors else "FAILURE"
        job.inserted = inserted; job.updated = updated; job.errors = errors
        job.finished_at = datetime.datetime.utcnow()
        s.commit()
        return {"inserted": inserted, "updated": updated, "errors": len(errors)}

    except Exception as e:
        s.rollback(); logger.exception(e)
        if 'job' in locals() and job:
            job.status = "FAILURE"
            job.errors = [{"error": str(e)}]
            job.finished_at = datetime.datetime.utcnow()
            s.commit()
        raise
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        s.close()

# -------------------------------------------------------------------
# import matching client / code client
# -------------------------------------------------------------------

@celery.task(name="task_import_customer_codes")
def task_import_customer_codes(job_id: int, tmp_path: str, labo_id: int = LABO_A_ID_DEFAULT, create_missing: bool = True):
    s = Session()
    try:
        job = s.get(ImportJob, job_id)
        job.task_id = task_import_customer_codes.request.id
        job.status = "STARTED"; s.commit()

        df = _read_any(tmp_path)
        # Normalisation noms de colonnes
        cols = { _norm(c): c for c in df.columns }
        def pick(*names):
            for n in names:
                if _norm(n) in cols: return cols[_norm(n)]
            for c in df.columns:
                if any(_norm(n) in _norm(c) for n in names):
                    return c
            return None

        c_code   = pick("code_client","code client","code","ref client")
        c_name   = pick("nom","name","client","enseigne","pharmacie")
        c_cp     = pick("cp","code_postal","zipcode")
        c_ville  = pick("ville")
        c_email  = pick("email","mail")
        c_phone  = pick("tel","telephone","téléphone","phone")

        if not (c_code and c_name):
            raise RuntimeError("Colonnes minimales manquantes : Code Client + nom.")

        # Renommage propre
        df = df.rename(columns={c_code:"code_client", c_name:"name"})
        if c_cp:    df = df.rename(columns={c_cp:"zipcode"})
        if c_ville: df = df.rename(columns={c_ville:"city"})
        if c_email: df = df.rename(columns={c_email:"email"})
        if c_phone: df = df.rename(columns={c_phone:"phone"})

        # Trim / normalisation
        for col in ["code_client","name","zipcode","city","email","phone"]:
            if col in df.columns:
                df[col] = df[col].astype(str).map(lambda x: " ".join(x.strip().split()))

        job.total_rows = int(len(df)); job.inserted = 0; job.updated = 0; job.errors = []; s.commit()
        inserted = updated = created_customers = 0
        unmatched = []

        # utilitaires
        customer_cols = _get_table_columns("customer")
        def norm_txt(x: str) -> str:
            import unicodedata, re
            x = (x or "").strip().lower()
            x = "".join(c for c in unicodedata.normalize("NFKD", x) if not unicodedata.combining(c))
            x = re.sub(r"\s+", " ", x)
            return x

        for i, r in df.iterrows():
            try:
                code_client = (r.get("code_client") or "").strip()
                name = (r.get("name") or "").strip()
                if not code_client or not name:
                    continue

                email = (r.get("email") or "").strip() if "email" in df.columns else None
                phone = (r.get("phone") or "").strip() if "phone" in df.columns else None
                zipcode = (r.get("zipcode") or "").strip() if "zipcode" in df.columns else None
                city = (r.get("city") or "").strip() if "city" in df.columns else None

                cust = None

                # 1) email
                if email and "email" in customer_cols:
                    cust = s.execute(select(Customer).where(func.lower(Customer.email) == func.lower(email))).scalar_one_or_none()

                # 2) phone
                if not cust and phone and "phone" in customer_cols:
                    cust = s.execute(select(Customer).where(func.lower(Customer.phone) == func.lower(phone))).scalar_one_or_none()

                # 3) name + zipcode + city
                if not cust and all([("name" in customer_cols), ("zipcode" in customer_cols), ("city" in customer_cols),
                                     zipcode, city]):
                    q = select(Customer).where(
                        func.lower(Customer.name) == func.lower(name),
                        func.lower(Customer.zipcode) == func.lower(zipcode),
                        func.lower(Customer.city) == func.lower(city),
                    ).limit(1)
                    cust = s.execute(q).scalar_one_or_none()

                # 4) name seul
                if not cust and "name" in customer_cols:
                    q = select(Customer).where(func.lower(Customer.name) == func.lower(name)).limit(1)
                    cust = s.execute(q).scalar_one_or_none()

                # Créer si toujours rien
                if not cust and create_missing and "name" in customer_cols:
                    cust = Customer(name=name)
                    if "zipcode" in customer_cols and zipcode:
                        setattr(cust, "zipcode", zipcode)
                    if "city" in customer_cols and city:
                        setattr(cust, "city", city)
                    if "address" in customer_cols and "ADRESSE" in df.columns:
                        setattr(cust, "address", str(r.get("ADRESSE") or "").strip())
                    if "email" in customer_cols and email:
                        setattr(cust, "email", email)
                    if "phone" in customer_cols and phone:
                        setattr(cust, "phone", phone)
                    if "labo_id" in customer_cols:
                        setattr(cust, "labo_id", labo_id)
                    s.add(cust); s.flush()
                    created_customers += 1

                if not cust:
                    unmatched.append({"row": int(i)+1, "name": name, "zipcode": zipcode, "city": city, "email": email, "phone": phone, "code_client": code_client})
                    if (i + 1) % 100 == 0:
                        job.errors = unmatched; s.commit()
                    continue

                # Upsert labo_client
                s.execute(
                    text("""
                        INSERT INTO labo_client (labo_id, client_id, code_client)
                        VALUES (:labo_id, :client_id, :code_client)
                        ON CONFLICT (labo_id, code_client)
                        DO UPDATE SET client_id = EXCLUDED.client_id
                    """),
                    {"labo_id": labo_id, "client_id": cust.id, "code_client": code_client}
                )
                s.flush()
                # On ne sait pas si c'est insert/update → on calcule naïf via existence préalable
                # Option: checker existence avant; ici on incrémente "updated" si conflit:
                # simplifions → on les compte tous en "updated" si déjà existait
                # Pour tracer correctement, on peut chercher avant:
                existing = s.execute(
                    text("SELECT 1 FROM labo_client WHERE labo_id=:l AND code_client=:c"),
                    {"l": labo_id, "c": code_client}
                ).first()
                if existing:
                    updated += 1
                else:
                    inserted += 1

                if (i + 1) % 200 == 0:
                    _save_job_progress(s, job, inserted, updated, unmatched)

            except Exception as e:
                s.rollback()
                unmatched.append({"row": int(i)+1, "error": str(e)})
                job.errors = unmatched; s.commit()

        job.status = "SUCCESS" if not unmatched else "FAILURE"
        job.inserted = inserted
        job.updated = updated
        job.errors = unmatched
        # on stocke aussi le compteur de créations clients dans 'updated' ou log
        job.finished_at = datetime.datetime.utcnow()
        s.commit()

        return {
            "inserted_links": inserted,
            "updated_links": updated,
            "created_customers": created_customers,
            "unmatched": len(unmatched),
        }

    except Exception as e:
        s.rollback()
        logger.exception(e)
        if 'job' in locals() and job:
            job.status = "FAILURE"
            job.errors = [{"error": str(e)}]
            job.finished_at = datetime.datetime.utcnow()
            s.commit()
        raise
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        s.close()




# -------------------------------------------------------------------
# Import COMMANDES “une ligne = un produit”
# -------------------------------------------------------------------
@celery.task(name="task_import_orders")
def task_import_orders(job_id: int, tmp_path: str, labo_id: int = LABO_A_ID_DEFAULT):
    s = Session()
    try:
        if Order is None or OrderItem is None:
            raise RuntimeError("Les modèles Order / OrderItem ne sont pas disponibles dans app.db.models.")

        job = s.get(ImportJob, job_id)
        job.task_id = task_import_orders.request.id
        job.status = "STARTED"; s.commit()

        df = _read_any(tmp_path)

        # Mapping souple des colonnes
        cols = { _norm(c): c for c in df.columns }
        def pick(*names):
            for n in names:
                n2 = _norm(n)
                if n2 in cols: return cols[n2]
            # fallback par inclusion
            for c in df.columns:
                if any(_norm(n) in _norm(c) for n in names):
                    return c
            return None

        c_order_number = pick("order_number","numero_commande","num_commande","num_piece","num pièce","ref_commande","commande")
        c_order_date   = pick("order_date","date","date_commande","date de commande")
        c_delivery_date = pick("delivery_date","date_livraison","deliv_date","ship_date")  # ✅ NOUVEAU
        c_client_name  = pick("client_name","nom_client","client","enseigne","pharmacie")
        c_client_code  = pick("code_client","client_code")
        c_status       = pick("status","statut")
        c_currency     = pick("currency","devise")
        c_payment      = pick("payment_method","mode_paiement","mode paiement")

        c_line_sku     = pick("line_sku","code_article","sku","reference","référence","ref")
        c_line_name    = pick("line_name","designation","désignation","libelle","libellé","nom_produit")
        c_line_qty     = pick("line_qty","quantite","quantité","qty","quantity")
        c_line_price   = pick("line_price_ht","pu_ht","prix_unitaire_ht","prix_ht","prix")
        c_line_total   = pick("line_total_ht","montant_total","total_ht_ligne")

        c_agent_name   = pick("agent_name","nom_representant","representant","commercial","vendeur")
        c_agent_email  = pick("agent_email","email_commercial","mail_commercial")

        if not (c_order_number and c_order_date and c_client_name and c_line_sku and c_line_qty):
            raise RuntimeError("Colonnes minimales manquantes : order_number, order_date, client_name, line_sku, line_qty.")

        # Normalisation des noms de colonnes
        rename_map = {
            c_order_number:"order_number",
            c_order_date:"order_date",
            c_client_name:"client_name",
            c_line_sku:"line_sku",
            (c_line_name or "line_name"):"line_name",
            c_line_qty:"line_qty",
        }
        if c_delivery_date: rename_map[c_delivery_date] = "delivery_date"    # ✅ NOUVEAU
        if c_line_price:    rename_map[c_line_price]    = "line_price_ht"
        if c_line_total:    rename_map[c_line_total]    = "line_total_ht"
        if c_client_code:   rename_map[c_client_code]   = "client_code"
        if c_status:        rename_map[c_status]        = "status"
        if c_currency:      rename_map[c_currency]      = "currency"
        if c_payment:       rename_map[c_payment]       = "payment_method"
        if c_agent_email:   rename_map[c_agent_email]   = "agent_email"
        if c_agent_name:    rename_map[c_agent_name]    = "agent_name"
        df = df.rename(columns=rename_map)

        # Typage
        df["order_date"] = df["order_date"].map(_parse_date_any)
        if "delivery_date" in df.columns:                                 # ✅ NOUVEAU
            df["delivery_date"] = df["delivery_date"].map(_parse_date_any)
        df["line_qty"] = pd.to_numeric(df["line_qty"], errors="coerce").fillna(0)
        if "line_price_ht" in df.columns:
            df["line_price_ht"] = df["line_price_ht"].map(_safe_float).fillna(0)
        if "line_total_ht" in df.columns:
            df["line_total_ht"] = df["line_total_ht"].map(_safe_float).fillna(pd.NA)

        # Calcul PU si seulement total fourni
        if "line_price_ht" not in df.columns and "line_total_ht" in df.columns:
            qty = df["line_qty"].replace(0, pd.NA)
            df["line_price_ht"] = (df["line_total_ht"] / qty).round(4).fillna(0)

        # bornage temporel 2 ans
        today = datetime.date.today()
        min_date = today - datetime.timedelta(days=370*2)
        df = df[(df["order_date"].notna()) & (df["order_date"] >= min_date)]

        job.total_rows = int(len(df)); job.inserted = 0; job.updated = 0; job.errors = []; s.commit()
        inserted = updated = 0
        errors: List[Dict[str, Any]] = []

        if len(df) == 0:
            job.status = "SUCCESS"; job.finished_at = datetime.datetime.utcnow(); s.commit()
            return {"inserted": 0, "updated": 0, "errors": 0}

        # Groupby par commande
        for order_number, g in df.groupby("order_number", dropna=True):
            try:
                head = g.iloc[0]
                order_date    = head.get("order_date")
                delivery_date = head.get("delivery_date") if "delivery_date" in g.columns else None  # ✅ NOUVEAU
                client_name   = _normalize_whitespace(head.get("client_name"))
                client_code   = _normalize_whitespace(head.get("client_code")) if "client_code" in g.columns else None
                status        = _normalize_whitespace(head.get("status")) if "status" in g.columns else None
                currency      = _normalize_whitespace(head.get("currency")) if "currency" in g.columns else "EUR"
                payment_method= _normalize_whitespace(head.get("payment_method")) if "payment_method" in g.columns else None

                # Résoudre agent
                agent_email = _normalize_whitespace(head.get("agent_email")) if "agent_email" in g.columns else None
                agent_name  = _normalize_whitespace(head.get("agent_name"))  if "agent_name"  in g.columns else None
                agent_obj   = _resolve_agent(s, agent_email, agent_name)
                agent_id    = agent_obj.id if agent_obj else None
                if agent_obj:
                    _ensure_agent_labo(s, agent_obj.id, labo_id)

                # Résoudre / créer customer
                cust = _find_or_create_customer(s, labo_id, client_name, client_code)
                customer_id = getattr(cust, "id", None) if cust else None

                # Lien labo ↔ client avec code client
                _ensure_labo_client(s, labo_id, cust, client_code)

                # Upsert entête Order (clé unique (labo_id, order_number))
                existing = s.execute(
                    select(Order).where(
                        getattr(Order, "labo_id") == labo_id,
                        getattr(Order, "order_number") == str(order_number).strip()
                    )
                ).scalar_one_or_none()

                data_header = {
                    "labo_id": labo_id,
                    "order_number": str(order_number).strip(),
                    "order_date": order_date,
                    "delivery_date": delivery_date,       # ✅ NOUVEAU
                    "status": status or "pending",
                    "currency": currency,
                    "payment_method": payment_method,
                    "agent_id": agent_id,
                    "customer_id": customer_id,
                    "client_name": client_name,
                }
                data_header = _only_existing_fields(data_header, "order")

                if existing:
                    for k, v in data_header.items():
                        setattr(existing, k, v if v is not None else getattr(existing, k))
                    o = existing
                    updated += 1
                else:
                    o = Order(**data_header)
                    s.add(o); s.flush()
                    inserted += 1

                # Remplace les items de la commande (idempotent)
                s.execute(text('DELETE FROM "order_item" WHERE order_id = :oid'), {"oid": o.id})

                calc_total_ht = 0.0
                for _, r in g.iterrows():
                    sku = str(r.get("line_sku") or "").strip()
                    if not sku:
                        continue

                    qty = float(r.get("line_qty") or 0)
                    # pht = prix unitaire HT (sera aussi utilisé pour unit_ht)
                    pht = float(r.get("line_price_ht") or 0)
                    if (not pht) and pd.notna(r.get("line_total_ht")) and qty:
                        pht = float(r.get("line_total_ht")) / qty

                    tht = round(qty * pht, 2)
                    calc_total_ht += tht

                    # Résoudre le produit pour obtenir product_id / ean13
                    prod = s.execute(
                        select(Product).where(Product.labo_id == labo_id, Product.sku == sku)
                    ).scalar_one_or_none()
                    product_id = prod.id if prod else None
                    ean13 = prod.ean13 if prod else None

                    # ✅ Alimente unit_ht (obligatoire en base) + price_ht si la colonne existe
                    # ✅ On alimente aussi la colonne line_ht, exigée par la base
                    item_data = {
                        "order_id": o.id,
                        "product_id": product_id,
                        "sku": sku,
                        "ean13": ean13,
                        "qty": qty,
                        "unit_ht": pht,       # Prix unitaire HT
                        "line_ht": tht,       # ✅ Nouveau : total de la ligne HT (obligatoire)
                        "price_ht": pht,      # Même valeur, filtrée si la colonne n'existe pas
                        "total_ht": tht,      # Idem
                    }
                    item_data = _only_existing_fields(item_data, "order_item")


                    # Si product_id requis et introuvable → on log l'erreur et on skip
                    if "product_id" in _get_table_columns("order_item") and item_data.get("product_id") is None:
                        errors.append({"order_number": str(order_number), "error": f"SKU inconnu pour ce labo: {sku}"})
                        job.errors = errors; s.commit()
                        continue

                    s.execute(
                        text(f'INSERT INTO "order_item" ({",".join(item_data.keys())}) '
                             f'VALUES ({",".join(":"+k for k in item_data.keys())})'),
                        item_data
                    )

                # Totaux en entête (si colonnes présentes)
                order_cols = _get_table_columns("order")
                if "total_ht" in order_cols:
                    setattr(o, "total_ht", round(calc_total_ht, 2))
                if "total_ttc" in order_cols:
                    setattr(o, "total_ttc", round(calc_total_ht, 2))  # sans TVA: TTC = HT

                if (inserted + updated) % 50 == 0:
                    _save_job_progress(s, job, inserted, updated, errors)

                s.commit()

            except Exception as e:
                s.rollback()
                errors.append({"order_number": str(order_number), "error": str(e)})
                job.errors = errors; s.commit()

        job.status = "SUCCESS" if not errors else "FAILURE"
        job.inserted = inserted; job.updated = updated; job.errors = errors
        job.finished_at = datetime.datetime.utcnow()
        s.commit()
        return {"inserted": inserted, "updated": updated, "errors": len(errors)}

    except Exception as e:
        s.rollback(); logger.exception(e)
        if 'job' in locals() and job:
            job.status = "FAILURE"
            job.errors = [{"error": str(e)}]
            job.finished_at = datetime.datetime.utcnow()
            s.commit()
        raise
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        s.close()


# -------------------------------------------------------------------
# PING simple
# -------------------------------------------------------------------
@celery.task(name="ping")
def ping():
    return "pong"
