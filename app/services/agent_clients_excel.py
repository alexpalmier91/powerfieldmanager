# app/services/agent_clients_excel.py
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple
from datetime import datetime, date
import re

from openpyxl import Workbook, load_workbook


# Colonnes du fichier Agent (pas de code_client : notion labo)
EXCEL_COLUMNS: List[str] = [
    "nom_societe",
    "contact_nom",
    "email",
    "telephone",
    "adresse",
    "code_postal",
    "ville",
    "pays",
    "siret",
    "tva_intracom",   # pas dans Client -> ignoré à l'import, None à l'export
    "actif",          # pas dans Client -> ignoré à l'import, TRUE à l'export
    "created_at",     # ignoré à l'import
]


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _excel_safe_value(v: Any) -> Any:
    """
    openpyxl refuse les datetimes timezone-aware.
    On convertit en naive (sans tzinfo).
    """
    if v is None:
        return None

    if isinstance(v, datetime):
        if v.tzinfo is not None:
            return v.replace(tzinfo=None)
        return v

    if isinstance(v, date):
        return v

    if isinstance(v, bool):
        return v

    # numbers / decimals etc -> ok
    return v


def build_clients_excel(rows: List[Dict[str, Any]]) -> BytesIO:
    """
    rows: liste de dicts avec les clés EXCEL_COLUMNS (ou superset)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    # Header
    ws.append(EXCEL_COLUMNS)

    # Rows
    for r in rows:
        ws.append([_excel_safe_value(r.get(col)) for col in EXCEL_COLUMNS])

    # Autosize simple
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        max_len = len(col_name)
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                if c.value is None:
                    continue
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _norm(s: str | None) -> str:
    if not s:
        return ""
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _to_bool(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        t = v.strip().lower()
        if t in ("true", "vrai", "1", "yes", "y", "oui"):
            return True
        if t in ("false", "faux", "0", "no", "n", "non"):
            return False
    return None


def parse_clients_excel(content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Tolère des colonnes en plus (ignorées).
    Attend au minimum les colonnes définies dans EXCEL_COLUMNS (sauf created_at optionnel).
    Retour:
      - parsed_rows : liste dict + _row_index + _invalid_email
      - errors : liste de messages (erreurs globales / header)
    """
    errors: List[str] = []
    rows_out: List[Dict[str, Any]] = []

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        return [], [f"Fichier Excel illisible: {exc}"]

    ws = wb.active

    # Header (ligne 1)
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_cells or not header_cells[0]:
        return [], ["Fichier Excel vide (aucune ligne d’en-tête)."]

    header = [str(h).strip() if h is not None else "" for h in header_cells[0]]
    header_map = {name: idx for idx, name in enumerate(header) if name}

    # Exigences minimales (created_at est ignoré à l'import, donc on le rend optionnel)
    required_cols = [c for c in EXCEL_COLUMNS if c not in ("created_at",)]
    missing = [c for c in required_cols if c not in header_map]
    if missing:
        return [], [f"Colonnes manquantes dans l’en-tête: {', '.join(missing)}"]

    # Parcours data (à partir de ligne 2)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # ignore lignes totalement vides
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        rec: Dict[str, Any] = {"_row_index": i, "_invalid_email": False}

        for col in EXCEL_COLUMNS:
            if col not in header_map:
                continue
            idx = header_map[col]
            val = row[idx] if idx < len(row) else None

            if isinstance(val, str):
                val = _norm(val)

            if col == "actif":
                val = _to_bool(val)

            rec[col] = val

        # Validation email si présent
        email = (rec.get("email") or "").strip()
        if email and not _EMAIL_RE.match(email):
            rec["_invalid_email"] = True

        # Trim simple pays/ville
        rec["pays"] = _norm(rec.get("pays"))
        rec["ville"] = _norm(rec.get("ville"))

        rows_out.append(rec)

    return rows_out, errors
